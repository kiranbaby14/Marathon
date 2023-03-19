import openpyxl
from bs4 import BeautifulSoup

# Open the file containing the WPT lines
with open("Lands_End_To_Fort_William_2023.gpx", "r") as f:
    data = f.read()

# Parse the XML data using BeautifulSoup and lxml parser
soup = BeautifulSoup(data, "lxml")

# Create a new Excel workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers for the latitude and longitude columns
worksheet.cell(row=1, column=1, value="Latitude")
worksheet.cell(row=1, column=2, value="Longitude")

# Find all the WPT lines
wpt_list = soup.find_all("wpt")

# Loop through each WPT line and extract the latitude and longitude information
row_number = 2  # start from row 2 to leave space for headers
for wpt in wpt_list:
    latitude = wpt["lat"]
    longitude = wpt["lon"]
    
    # Write the latitude and longitude values to the Excel worksheet
    worksheet.cell(row=row_number, column=1, value=latitude)
    worksheet.cell(row=row_number, column=2, value=longitude)
    
    row_number += 1

# Save the Excel workbook
workbook.save("output_file.xlsx")
